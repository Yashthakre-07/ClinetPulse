import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
import pandas as pd
import sqlite3
import os

DB_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), "client_pulse.db")
OUTPUT_EXCEL = os.path.join(os.path.dirname(os.path.dirname(__file__)), "ClientPulse_Financial_Report.xlsx")

def generate_styled_excel_report():
    """Generates an executive-ready, styled Excel model using openpyxl."""
    conn = sqlite3.connect(DB_PATH)
    latest_month_query = "SELECT MAX(ym_month) FROM fact_monthly_financials;"
    latest_month = pd.read_sql_query(latest_month_query, conn).iloc[0, 0]

    df = pd.read_sql_query(
        f"SELECT client_id, client_name, industry, region, contract_tier, gross_revenue, total_cost, net_profit, profit_margin_pct, mom_growth_pct, rolling_3m_avg_revenue, avg_csat, churn_risk_flag FROM fact_monthly_financials WHERE ym_month = '{latest_month}' ORDER BY gross_revenue DESC;",
        conn
    )
    conn.close()

    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executive Summary"
    ws.views.sheetView[0].showGridLines = True

    # Styling Palettes
    HEADER_FILL = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Navy
    CARD_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Soft Gray
    ALERT_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Soft Red/Orange
    
    WHITE_BOLD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F497D")
    SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="595959")
    CARD_HEADER_FONT = Font(name="Calibri", size=9, bold=True, color="595959")
    CARD_VAL_FONT = Font(name="Calibri", size=14, bold=True, color="1F497D")
    DATA_FONT = Font(name="Calibri", size=10)

    THIN_BORDER = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # 1. Title Banner
    ws['A1'] = "ClientPulse — Executive Financial & Analytics Report"
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f"Automated Reporting Model | Period: {latest_month} | Confidential"
    ws['A2'].font = SUBTITLE_FONT

    # 2. Executive KPI Summary Cards (Row 4-6)
    kpis = [
        ("TOTAL MONTHLY REVENUE", f"=SUM(F9:F{8+len(df)})", "$#,##0"),
        ("ANNUAL RUN RATE (ARR)", f"=F4*12", "$#,##0"),
        ("AVG PROFIT MARGIN", f"=AVERAGE(I9:I{8+len(df)})/100", "0.0%"),
        ("ACTIVE CLIENTS", len(df), "0"),
        ("CHURN RISK ACCOUNTS", f"=COUNTIF(M9:M{8+len(df)}, 1)", "0")
    ]

    col_idx = 1
    for label, val_formula, num_fmt in kpis:
        col_letter1 = get_column_letter(col_idx)
        col_letter2 = get_column_letter(col_idx + 1) if col_idx < 9 else col_letter1
        
        ws[f'{col_letter1}4'] = label
        ws[f'{col_letter1}4'].font = CARD_HEADER_FONT
        ws[f'{col_letter1}4'].fill = CARD_FILL
        ws[f'{col_letter1}4'].alignment = Alignment(horizontal="center")

        ws[f'{col_letter1}5'] = val_formula
        ws[f'{col_letter1}5'].font = CARD_VAL_FONT
        ws[f'{col_letter1}5'].fill = CARD_FILL
        ws[f'{col_letter1}5'].number_format = num_fmt
        ws[f'{col_letter1}5'].alignment = Alignment(horizontal="center")

        ws[f'{col_letter1}4'].border = THIN_BORDER
        ws[f'{col_letter1}5'].border = THIN_BORDER
        col_idx += 2

    # 3. Data Table Headers (Row 8)
    headers = [
        "Client ID", "Client Name", "Industry", "Region", "Contract Tier",
        "Gross Revenue", "Total Cost", "Net Profit", "Profit Margin %",
        "MoM Growth %", "3M Avg Revenue", "Avg CSAT", "Churn Risk"
    ]

    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=8, column=c_idx, value=h)
        cell.font = WHITE_BOLD
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center" if c_idx not in [2, 3, 4] else "left", vertical="center")

    # 4. Populate Data Rows
    for r_idx, row in df.iterrows():
        row_num = 8 + r_idx + 1
        
        values = [
            row['client_id'], row['client_name'], row['industry'], row['region'], row['contract_tier'],
            row['gross_revenue'], row['total_cost'], row['net_profit'], row['profit_margin_pct'] / 100.0,
            row['mom_growth_pct'] / 100.0, row['rolling_3m_avg_revenue'], row['avg_csat'],
            "RISK" if row['churn_risk_flag'] == 1 else "STABLE"
        ]

        for c_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            # Formatting
            if c_idx in [6, 7, 8, 11]: # Currency
                cell.number_format = "$#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif c_idx in [9, 10]: # Percentage
                cell.number_format = "0.0%"
                cell.alignment = Alignment(horizontal="right")
            elif c_idx == 12: # CSAT
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="center")
            elif c_idx == 13: # Churn Risk Flag
                cell.alignment = Alignment(horizontal="center")
                if val == "RISK":
                    cell.fill = ALERT_FILL
                    cell.font = Font(name="Calibri", size=10, bold=True, color="C00000")
            else:
                cell.alignment = Alignment(horizontal="left" if c_idx in [2, 3, 4] else "center")

    # 5. Embed Bar Chart for Top Client Revenues
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Client Revenue Breakdown ($)"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Client Name"

    data_ref = Reference(ws, min_col=6, min_row=8, max_row=8 + len(df))
    cats_ref = Reference(ws, min_col=2, min_row=9, max_row=8 + len(df))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.height = 12
    chart.width = 18
    ws.add_chart(chart, f"B{11 + len(df)}")

    # 6. Auto-fit Columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(OUTPUT_EXCEL)
    print(f"[EXCEL EXPORTER] Generated financial report: '{OUTPUT_EXCEL}' successfully!")
    return OUTPUT_EXCEL

if __name__ == "__main__":
    generate_styled_excel_report()
